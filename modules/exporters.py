"""
Export functions for statistics in various formats (CSV, XLSX, JSON).
"""

import csv
import json
from pathlib import Path
from datetime import datetime
from colorama import Fore, Style
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


def export_statistics_csv(stats: dict, output_path: Path, total_files: int, files_with_hits: int, total_matches: int):
    """Export statistics to CSV format."""
    try:
        with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)

            # Write header
            writer.writerow(['Category', 'Sensitive Term', 'Total Occurrences'])

            # Write data
            for category, words in sorted(stats.items()):
                for word, count in sorted(words.items(), key=lambda x: x[1], reverse=True):
                    writer.writerow([category, word, count])

            # Write summary
            writer.writerow([])
            writer.writerow(['Summary Statistics'])
            writer.writerow(['Total files scanned', total_files])
            writer.writerow(['Files with hits', files_with_hits])
            writer.writerow(['Total matches found', total_matches])
            writer.writerow(['Unique terms found', sum(len(words) for words in stats.values())])
            writer.writerow(['Categories with findings', len(stats)])

        print(f"{Fore.GREEN}[OK] CSV report exported: {Style.BRIGHT}{output_path}{Style.RESET_ALL}")
    except Exception as e:
        print(f"{Fore.RED}[ERROR] Error exporting CSV: {e}{Style.RESET_ALL}")


def export_statistics_xlsx(stats: dict, output_path: Path, total_files: int, files_with_hits: int, total_matches: int):
    """Export statistics to XLSX format with formatting in two sheets."""
    try:
        wb = openpyxl.Workbook()

        # Header formatting
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        # ========== Sheet 1: Occurrences ==========
        ws_occurrences = wb.active
        ws_occurrences.title = "Occurrences"

        # Write headers for Occurrences sheet
        headers = ['Category', 'Sensitive Term', 'Total Occurrences']
        for col, header in enumerate(headers, 1):
            cell = ws_occurrences.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Write data
        row = 2
        for category, words in sorted(stats.items()):
            for word, count in sorted(words.items(), key=lambda x: x[1], reverse=True):
                ws_occurrences.cell(row=row, column=1, value=category)
                ws_occurrences.cell(row=row, column=2, value=word)
                ws_occurrences.cell(row=row, column=3, value=count)
                row += 1

        # Adjust column widths
        ws_occurrences.column_dimensions['A'].width = 25
        ws_occurrences.column_dimensions['B'].width = 30
        ws_occurrences.column_dimensions['C'].width = 20

        # ========== Sheet 2: Statistics ==========
        ws_stats = wb.create_sheet(title="Statistics")

        # Summary formatting
        summary_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        summary_font = Font(bold=True)
        label_font = Font(bold=True)

        # Write summary statistics
        row = 1
        summary_data = [
            ['Metric', 'Value'],
            ['Total files scanned', total_files],
            ['Files with hits', files_with_hits],
            ['Total matches found', total_matches],
            ['Unique terms found', sum(len(words) for words in stats.values())],
            ['Categories with findings', len(stats)]
        ]

        for idx, (label, value) in enumerate(summary_data):
            cell_label = ws_stats.cell(row=row + idx, column=1, value=label)
            cell_value = ws_stats.cell(row=row + idx, column=2, value=value)

            if idx == 0:  # Header row
                cell_label.fill = header_fill
                cell_label.font = header_font
                cell_label.alignment = Alignment(horizontal='center')
                cell_value.fill = header_fill
                cell_value.font = header_font
                cell_value.alignment = Alignment(horizontal='center')
            else:
                cell_label.font = label_font

        # Adjust column widths
        ws_stats.column_dimensions['A'].width = 30
        ws_stats.column_dimensions['B'].width = 20

        wb.save(output_path)
        print(f"{Fore.GREEN}[OK] XLSX report exported: {Style.BRIGHT}{output_path}{Style.RESET_ALL}")
    except Exception as e:
        print(f"{Fore.RED}[ERROR] Error exporting XLSX: {e}{Style.RESET_ALL}")


def export_statistics_json(stats: dict, output_path: Path, total_files: int, files_with_hits: int, total_matches: int, file_types: dict):
    """Export statistics to JSON format."""
    try:
        # Prepare data structure
        export_data = {
            "summary": {
                "total_files_scanned": total_files,
                "files_with_hits": files_with_hits,
                "total_matches": total_matches,
                "unique_terms_found": sum(len(words) for words in stats.values()),
                "categories_with_findings": len(stats),
                "scan_timestamp": datetime.now().isoformat()
            },
            "file_types": file_types,
            "findings": {}
        }

        # Convert stats to JSON-friendly format
        for category, words in sorted(stats.items()):
            export_data["findings"][category] = []
            for word, count in sorted(words.items(), key=lambda x: x[1], reverse=True):
                export_data["findings"][category].append({
                    "term": word,
                    "occurrences": count
                })

        with open(output_path, 'w', encoding='utf-8') as jsonfile:
            json.dump(export_data, jsonfile, indent=2, ensure_ascii=False)

        print(f"{Fore.GREEN}[OK] JSON report exported: {Style.BRIGHT}{output_path}{Style.RESET_ALL}")
    except Exception as e:
        print(f"{Fore.RED}[ERROR] Error exporting JSON: {e}{Style.RESET_ALL}")
